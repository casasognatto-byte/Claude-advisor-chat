"""Catálogo de cores de MDP/MDF das linhas que o Grupo Simonetto trabalha —
Simonetto e Stimmo. Importado de uma planilha .xlsx que o Davi mantém e
reenvia de tempos em tempos (painel do diretor tem o botão de reimportar).

A planilha real observada (`CORES DAS LINHAS.xlsx`) tem uma aba por ano
(2023, 2024, 2025...) com layout: linha de cabeçalho com o nome de cada marca
("SIMONETTO", "STIMMO"/"ESTIMMO") numa coluna, e abaixo dela pares
(nome da cor, fabricante da placa) até a lista acabar. Pegamos sempre a aba
mais recente (maior ano no nome) e localizamos as colunas pelo texto do
cabeçalho, não pela posição fixa — para tolerar pequenas mudanças de layout
entre uma planilha e outra.

Imports de `app.main` ficam dentro das funções para evitar import circular.
"""

import io
import os
import re
import secrets
import unicodedata

from fastapi import APIRouter, File, HTTPException, Request, UploadFile

router = APIRouter(prefix="/api/materials")

# Aliases conhecidos pro nome de cada marca no cabeçalho da planilha —
# "ESTIMMO" apareceu como erro de digitação de "STIMMO" numa das abas.
BRAND_ALIASES = {
    "simonetto": "simonetto",
    "stimmo": "stimmo",
    "estimmo": "stimmo",
}
BRAND_LABELS = {"simonetto": "Simonetto", "stimmo": "Stimmo"}

# --- Swatches (imagem real do material) -------------------------------------
# 14/07/2026: pedido do Davi — o botão "Cores" só inseria o NOME da cor no
# prompt, sem precisão nenhuma pro Nano Banana (ele "chuta" a cor a partir só
# do texto). Recortamos manualmente o swatch real de cada cor a partir dos
# catálogos oficiais dos fabricantes (Arauco, Duratex, Berneck, Green Plac —
# Guararapes não tem swatch isolado, só foto de ambiente, então fica sem
# imagem) e guardamos em app/static/color_swatches/. Esse arquivo liga o nome
# da cor (como aparece na planilha do Davi) ao arquivo do swatch.
#
# _swatch_core_name() remove sufixos de acabamento/variação (CHESS, MATT, TX,
# "- AF" etc.) que a planilha às vezes anexa ao nome — sem isso "CANELA
# CHESS" não bateria com o swatch "CANELA" que recortamos.
_SWATCH_SUFFIXES = (
    "CHESS", "MATT", "TX", "AF", "VEL", "MICRO", "RUST", "DESIGN", "LSF",
    "ALUMI", "TEXTURA",
)

SWATCH_DIR = os.path.join(os.path.dirname(__file__), "static", "color_swatches")

# nome-nucleo (ver _swatch_core_name) -> arquivo em app/static/color_swatches/
SWATCH_FILES = {
    # Arauco
    "ACACIA CARMEL": "ACACIA_CARMEL.png", "ATENNA": "ATENNA.png", "BEIGE": "BEIGE.png",
    "BRANCO SUPREMO": "BRANCO_SUPREMO.png", "BRANCO": "BRANCO_SUPREMO.png", "CANELA": "CANELA.png",
    "CINZA CRISTAL": "CINZA_CRISTAL.png", "CINZA PURO": "CINZA_PURO.png",
    "CONCRETO DECOR": "CONCRETO_DECOR.png", "CONNECT": "CONNECT.png", "EBANO": "EBANO.png",
    "GRAFITO": "GRAFITO.png", "JADE": "JADE.png", "LINHO": "LINHO.png",
    "LINO PIOMBO": "LINO_PIOMBO.png", "LINO PIMBO": "LINO_PIOMBO.png", "LORD": "LORD.png",
    "LOURO": "LOURO.png", "NOCE NATURALE": "NOCE_NATURALE.png", "NOGUEIRA PERSA": "NOGUEIRA_PERSA.png",
    "AREAL": "AREAL.png", "BETON": "BETON.png", "CACAO": "CACAO.png", "CRISTALINA": "CRISTALINA.png",
    "DAMASCO": "DAMASCO.png", "ESCARLATE": "ESCARLATE.png", "FRAPE": "FRAPE.png", "GRIS": "GRIS.png",
    "PAU FERRO": "PAU_FERRO.png", "SAL ROSA": "SAL_ROSA.png", "SALVIA": "SALVIA.png",
    # Duratex
    "ABSOLUTO": "ABSOLUTO.png", "BLUSH": "BLUSH.png", "CINZA SAGRADO": "CINZA_SAGRADO.png",
    "GIANDUIA": "GIANDUIA.png", "NAZCA": "NAZCA.png", "NOGUEIRA CAIENA": "NOGUEIRA_CAIENA.png",
    "NOGUEIRA FLORIDA": "NOGUEIRA_FLORIDA.png", "PINOLE": "PINOLE.png", "TITANIO": "TITANIO.png",
    "PRETO": "PRETO.png",
    # Berneck
    "TABASCO": "TABASCO.png", "FALESIA": "FALESIA.png", "CERAMIK": "CERAMIK.png",
    "METALIC SUEDE": "METALIC_SUEDE.png", "METTALIC SUEDE": "METALIC_SUEDE.png",
    # Green Plac
    "LONDRES": "LONDRES.png", "VEREDAS": "VEREDAS.png", "NILO": "NILO.png",
    # Arauco — fotos de produto direto de arauco.com.br/lojaonline.arauco.com
    # (links que o Davi mandou, 14/07/2026), pras cores que não apareciam nos
    # catálogos em PDF.
    "VERMONT DARK": "VERMONT_DARK.jpg", "CARAMELO": "CAMELO.jpg", "BUFALO": "BUFALO.jpg",
    "TOKAI": "TOKAI.jpg", "NORDIC": "NORDIC.jpg",
}


def _swatch_core_name(name: str) -> str:
    s = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9 ]", " ", s.upper())
    tokens = [t for t in s.split() if t not in _SWATCH_SUFFIXES]
    return " ".join(tokens).strip()


def _swatch_file_for(name: str) -> str | None:
    return SWATCH_FILES.get(_swatch_core_name(name))


def init_materials_db() -> None:
    from app.main import DB_ENABLED, _db

    if not DB_ENABLED:
        return
    try:
        with _db() as conn, conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS material_colors (
                    id           TEXT PRIMARY KEY,
                    brand        TEXT NOT NULL,
                    name         TEXT NOT NULL,
                    manufacturer TEXT,
                    imported_by  TEXT NOT NULL,
                    imported_at  TIMESTAMPTZ NOT NULL DEFAULT now()
                );
                """
            )
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_material_colors_brand "
                "ON material_colors (brand, name);"
            )
            # 14/07/2026 — swatch real do material (ver SWATCH_FILES acima),
            # caminho relativo dentro de app/static/color_swatches/.
            cur.execute("ALTER TABLE material_colors ADD COLUMN IF NOT EXISTS swatch_file TEXT;")
            # Reconecta os swatches a cada boot pras linhas já importadas antes
            # deste recurso existir (produção já tinha 74 cores da planilha
            # antiga) — idempotente, sem precisar o Davi reimportar nada.
            cur.execute("SELECT id, name FROM material_colors WHERE swatch_file IS NULL")
            for color_id, name in cur.fetchall():
                swatch_file = _swatch_file_for(name)
                if swatch_file:
                    cur.execute(
                        "UPDATE material_colors SET swatch_file = %s WHERE id = %s",
                        (swatch_file, color_id),
                    )
    except Exception as e:
        print(f"[init_materials_db] falha: {e}")


def _parse_workbook(content: bytes) -> dict:
    """Devolve {"simonetto": [(nome, fabricante), ...], "stimmo": [...]}.
    Lança ValueError com mensagem clara se não achar nenhuma marca conhecida."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    # Prefere a aba com o maior ano no nome (planilha real tem uma aba por
    # ano); se nenhuma aba parecer um ano, usa a última (ordem de criação).
    year_sheets = [(int(m.group()), name) for name in wb.sheetnames if (m := re.search(r"\d{4}", name))]
    sheet_name = max(year_sheets)[1] if year_sheets else wb.sheetnames[-1]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Aba '{sheet_name}' está vazia.")
    header = rows[0]

    brand_cols: dict[str, int] = {}
    for col_idx, cell in enumerate(header):
        if not isinstance(cell, str):
            continue
        key = cell.strip().lower()
        if key in BRAND_ALIASES:
            brand_cols[BRAND_ALIASES[key]] = col_idx

    if not brand_cols:
        raise ValueError(
            f"Não encontrei nenhuma coluna 'SIMONETTO' ou 'STIMMO' no cabeçalho da aba '{sheet_name}'."
        )

    result: dict[str, list] = {brand: [] for brand in brand_cols}
    for row in rows[1:]:
        for brand, col_idx in brand_cols.items():
            if col_idx >= len(row):
                continue
            name = row[col_idx]
            if not isinstance(name, str) or not name.strip():
                continue
            manufacturer = None
            if col_idx + 1 < len(row) and isinstance(row[col_idx + 1], str) and row[col_idx + 1].strip():
                manufacturer = row[col_idx + 1].strip()
            result[brand].append((name.strip(), manufacturer))

    return result


@router.post("/import")
async def import_colors(request: Request, file: UploadFile = File(...)):
    """Substitui o catálogo inteiro pelo conteúdo da planilha enviada — só o
    diretor pode (mesma régua de outras ações administrativas)."""
    from app.main import _db, _require_db, require_admin

    user = require_admin(request)
    _require_db()
    content = await file.read()
    try:
        parsed = _parse_workbook(content)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Não consegui ler a planilha: {e}")

    total = sum(len(v) for v in parsed.values())
    if not total:
        raise HTTPException(400, "Planilha lida, mas nenhuma cor encontrada nas colunas de marca.")

    with _db() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM material_colors")
        for brand, colors in parsed.items():
            for name, manufacturer in colors:
                cur.execute(
                    "INSERT INTO material_colors (id, brand, name, manufacturer, imported_by, swatch_file) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    ("mc" + secrets.token_hex(8), brand, name, manufacturer, user["username"], _swatch_file_for(name)),
                )

    return {"ok": True, "counts": {brand: len(colors) for brand, colors in parsed.items()}}


@router.post("/colors/relink-swatches")
def relink_swatches(request: Request):
    """Reprocessa SWATCH_FILES contra as cores já importadas, sem precisar
    reimportar a planilha inteira — útil depois de adicionar/trocar recortes
    de swatch no código (deploy novo) sem esperar o Davi reenviar o .xlsx."""
    from app.main import _db, _require_db, require_admin

    require_admin(request)
    _require_db()
    with _db() as conn, conn.cursor() as cur:
        cur.execute("SELECT id, name FROM material_colors")
        rows = cur.fetchall()
        linked = 0
        for color_id, name in rows:
            swatch_file = _swatch_file_for(name)
            if swatch_file:
                linked += 1
            cur.execute(
                "UPDATE material_colors SET swatch_file = %s WHERE id = %s",
                (swatch_file, color_id),
            )
    return {"ok": True, "total": len(rows), "linked": linked}


@router.get("/colors")
def list_colors(request: Request, brand: str | None = None):
    from app.main import DB_ENABLED, _db, require_user

    require_user(request)
    if not DB_ENABLED:
        return []
    with _db() as conn, conn.cursor() as cur:
        if brand:
            cur.execute(
                "SELECT id, brand, name, manufacturer, swatch_file FROM material_colors "
                "WHERE brand = %s ORDER BY name",
                (brand,),
            )
        else:
            cur.execute(
                "SELECT id, brand, name, manufacturer, swatch_file FROM material_colors "
                "ORDER BY brand, name"
            )
        rows = cur.fetchall()
    return [
        {
            "id": r[0], "brand": r[1], "brandLabel": BRAND_LABELS.get(r[1], r[1]), "name": r[2],
            "manufacturer": r[3],
            "swatchUrl": f"/static/color_swatches/{r[4]}" if r[4] else None,
        }
        for r in rows
    ]


def get_swatches(color_ids: list[str], targets: dict[str, str] | None = None) -> list[dict]:
    """Usado por app/image.py na hora de gerar o render — devolve os bytes do
    swatch (imagem real do material) de cada cor selecionada, pra anexar como
    referência visual junto com a imagem elementar. Ids sem swatch_file (ex:
    cor da Guararapes, sem swatch isolado) ou inválidos são ignorados.

    `targets` (opcional) é {color_id: "nome do móvel"} — quando presente, vai
    junto no dict de retorno pra app/image_engines.py incluir na instrução de
    cada referência qual móvel deve receber aquela cor (ambientes com mais de
    um móvel, ver pedido do Davi de 14/07/2026)."""
    from app.main import DB_ENABLED, _db

    if not DB_ENABLED or not color_ids:
        return []
    targets = targets or {}
    with _db() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT id, name, brand, swatch_file FROM material_colors "
            "WHERE id = ANY(%s) AND swatch_file IS NOT NULL",
            (color_ids,),
        )
        rows = cur.fetchall()
    out = []
    for color_id, name, brand, swatch_file in rows:
        path = os.path.join(SWATCH_DIR, swatch_file)
        try:
            with open(path, "rb") as f:
                data = f.read()
        except OSError:
            continue
        mime = "image/jpeg" if swatch_file.lower().endswith((".jpg", ".jpeg")) else "image/png"
        out.append({
            "name": name, "brand": BRAND_LABELS.get(brand, brand), "bytes": data, "mime": mime,
            "target": targets.get(color_id) or None,
        })
    return out


@router.get("/status")
def import_status(request: Request):
    """Pro painel admin mostrar quando foi a última importação e quantas
    cores tem hoje, sem precisar listar tudo."""
    from app.main import DB_ENABLED, _db, require_admin

    require_admin(request)
    if not DB_ENABLED:
        return {"total": 0, "byBrand": {}, "lastImportedAt": None, "lastImportedBy": None}
    with _db() as conn, conn.cursor() as cur:
        cur.execute("SELECT brand, COUNT(*) FROM material_colors GROUP BY brand")
        by_brand = {r[0]: r[1] for r in cur.fetchall()}
        cur.execute(
            "SELECT imported_by, EXTRACT(EPOCH FROM imported_at) FROM material_colors "
            "ORDER BY imported_at DESC LIMIT 1"
        )
        last = cur.fetchone()
    return {
        "total": sum(by_brand.values()),
        "byBrand": by_brand,
        "lastImportedAt": float(last[1]) if last else None,
        "lastImportedBy": last[0] if last else None,
    }
